import io
from datetime import date

from openpyxl import load_workbook

from app.models import AttendanceDaily, Department
from app.models.enums import AttendanceStatus, UserRole


async def _login(client, employee, password) -> str:
    response = await client.post(
        "/v1/auth/login", json={"official_email": employee.official_email, "password": password}
    )
    assert response.status_code == 200
    return response.json()["access_token"]


def _auth_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


async def test_employee_and_manager_forbidden(client, make_employee):
    for role in (UserRole.EMPLOYEE, UserRole.MANAGER):
        employee, _, password = await make_employee(role=role)
        token = await _login(client, employee, password)
        response = await client.get(
            "/v1/reports/attendance-summary?month=2026-09", headers=_auth_headers(token)
        )
        assert response.status_code == 403


async def test_hr_gets_json_summary_with_correct_counts(client, make_employee, db_session):
    hr, _, hr_password = await make_employee(role=UserRole.HR)
    employee, _, _ = await make_employee(role=UserRole.EMPLOYEE)
    hr_token = await _login(client, hr, hr_password)

    db_session.add_all(
        [
            AttendanceDaily(
                employee_id=employee.id, work_date=date(2026, 9, 1), status=AttendanceStatus.PRESENT,
                worked_minutes=480, overtime_minutes=0, late_by_minutes=0,
            ),
            AttendanceDaily(
                employee_id=employee.id, work_date=date(2026, 9, 2), status=AttendanceStatus.ABSENT,
            ),
            AttendanceDaily(
                employee_id=employee.id, work_date=date(2026, 9, 3), status=AttendanceStatus.PRESENT,
                worked_minutes=500, overtime_minutes=20, late_by_minutes=5,
            ),
        ]
    )
    await db_session.flush()

    response = await client.get(
        "/v1/reports/attendance-summary?month=2026-09", headers=_auth_headers(hr_token)
    )
    assert response.status_code == 200
    row = next(r for r in response.json() if r["employee_id"] == employee.id)
    assert row["present"] == 2
    assert row["absent"] == 1
    assert row["worked_minutes"] == 980
    assert row["overtime_minutes"] == 20
    assert row["late_by_minutes"] == 5


async def test_department_filter_excludes_other_departments(client, make_employee, db_session):
    hr, _, hr_password = await make_employee(role=UserRole.HR)
    dept = Department(name="Engineering Test", code=f"ENGT-{id(object())}", is_active=True)
    db_session.add(dept)
    await db_session.flush()

    in_dept, _, _ = await make_employee(role=UserRole.EMPLOYEE, department_id=dept.id)
    out_of_dept, _, _ = await make_employee(role=UserRole.EMPLOYEE)
    hr_token = await _login(client, hr, hr_password)

    response = await client.get(
        f"/v1/reports/attendance-summary?month=2026-09&department={dept.id}", headers=_auth_headers(hr_token)
    )
    assert response.status_code == 200
    employee_ids = {r["employee_id"] for r in response.json()}
    assert in_dept.id in employee_ids
    assert out_of_dept.id not in employee_ids


async def test_xlsx_export_is_a_valid_workbook(client, make_employee, db_session):
    hr, _, hr_password = await make_employee(role=UserRole.HR)
    employee, _, _ = await make_employee(role=UserRole.EMPLOYEE)
    hr_token = await _login(client, hr, hr_password)

    db_session.add(
        AttendanceDaily(
            employee_id=employee.id, work_date=date(2026, 9, 1), status=AttendanceStatus.PRESENT,
            worked_minutes=480,
        )
    )
    await db_session.flush()

    response = await client.get(
        "/v1/reports/attendance-summary?month=2026-09&format=xlsx", headers=_auth_headers(hr_token)
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert header_row[0] == "Employee Code"
    assert sheet.max_row >= 2
